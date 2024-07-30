import openpyxl
from openpyxl.styles import Font
import tkinter as tk
from tkinter import messagebox


# Função para criar uma nova planilha de livro caixa
def criar_livro_caixa(nome_arquivo):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Livro Caixa"

    # Cabeçalhos
    headers = ["Data", "Descrição", "Entrada", "Saída", "Saldo"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    workbook.save(nome_arquivo)


# Função para adicionar um registro ao livro caixa
def adicionar_registro(nome_arquivo, data, descricao, entrada, saida):
    workbook = openpyxl.load_workbook(nome_arquivo)
    sheet = workbook["Livro Caixa"]

    # Encontrar a próxima linha vazia
    next_row = sheet.max_row + 1

    # Calcular o saldo
    if next_row == 2:  # Se for a primeira linha de dados, não há saldo anterior
        saldo_anterior = 0.0
    else:
        saldo_anterior = sheet.cell(row=next_row - 1, column=5).value
        if saldo_anterior is None:
            saldo_anterior = 0.0
        else:
            saldo_anterior = float(saldo_anterior)

    saldo_atual = saldo_anterior + entrada - saida

    # Adicionar os valores na planilha
    sheet.cell(row=next_row, column=1, value=data)
    sheet.cell(row=next_row, column=2, value=descricao)
    sheet.cell(row=next_row, column=3, value=entrada)
    sheet.cell(row=next_row, column=4, value=saida)
    sheet.cell(row=next_row, column=5, value=saldo_atual)

    workbook.save(nome_arquivo)


# Função para obter os valores dos campos de entrada e adicionar um registro
def obter_valores_e_adicionar():
    data = entry_data.get()
    descricao = entry_descricao.get()
    try:
        entrada = float(entry_entrada.get())
        saida = float(entry_saida.get())
    except ValueError:
        messagebox.showerror("Erro de valor", "Por favor, insira valores numéricos válidos para entrada e saída.")
        return

    adicionar_registro(nome_arquivo, data, descricao, entrada, saida)
    messagebox.showinfo("Registro adicionado", "Registro adicionado com sucesso!")
    limpar_campos()


# Função para limpar os campos de entrada
def limpar_campos():
    entry_data.delete(0, tk.END)
    entry_descricao.delete(0, tk.END)
    entry_entrada.delete(0, tk.END)
    entry_saida.delete(0, tk.END)


# Criar a interface gráfica
root = tk.Tk()
root.title("Livro Caixa")

# Criar os widgets
label_data = tk.Label(root, text="Data (AAAA-MM-DD):")
label_data.grid(row=0, column=0, padx=10, pady=5)

entry_data = tk.Entry(root)
entry_data.grid(row=0, column=1, padx=10, pady=5)

label_descricao = tk.Label(root, text="Descrição:")
label_descricao.grid(row=1, column=0, padx=10, pady=5)

entry_descricao = tk.Entry(root)
entry_descricao.grid(row=1, column=1, padx=10, pady=5)

label_entrada = tk.Label(root, text="Entrada:")
label_entrada.grid(row=2, column=0, padx=10, pady=5)

entry_entrada = tk.Entry(root)
entry_entrada.grid(row=2, column=1, padx=10, pady=5)

label_saida = tk.Label(root, text="Saída:")
label_saida.grid(row=3, column=0, padx=10, pady=5)

entry_saida = tk.Entry(root)
entry_saida.grid(row=3, column=1, padx=10, pady=5)

button_adicionar = tk.Button(root, text="Adicionar Registro", command=obter_valores_e_adicionar)
button_adicionar.grid(row=4, column=0, columnspan=2, padx=10, pady=10)

# Criar o arquivo de livro caixa
nome_arquivo = "C:/Users/matheus.araujo/Desktop/teste_planilha.xlsx"
criar_livro_caixa(nome_arquivo)

# Iniciar o loop principal da interface gráfica
root.mainloop()
